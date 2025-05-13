from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from models.usuario import Usuarios
from models.facturacion import Factura
from models.reserva import Reserva
from models.pqrs import PQRS
from models.casas import Apartamento, Torre, Casas
from models.roles import Roles
from models.datos_conjunto import DatosConjunto
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import os
from app import db
from datetime import datetime
from controllers.log import registrar_log
from controllers.configuracion import require_permission
from sqlalchemy import and_
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment

reportes_bp = Blueprint('reportes', __name__)

def generar_pdf(titulo, datos, columnas, filename, usuario):
    carpeta_reportes = os.path.abspath(os.path.join("static", "reportes"))
    os.makedirs(carpeta_reportes, exist_ok=True)

    filepath = os.path.join(carpeta_reportes, filename)
    c = canvas.Canvas(filepath, pagesize=landscape(letter))
    width, height = landscape(letter)

    datos_conjunto = DatosConjunto.query.first()
    nombre_conjunto = datos_conjunto.nombre if datos_conjunto else "Conjunto Residencial"

    logo_path = os.path.abspath(os.path.join("static", "img", "logo.png"))
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 50, height - 100, width=120, height=80, preserveAspectRatio=True, mask='auto')

    c.setFont("Helvetica-Bold", 16)
    c.drawString(220, height - 50, nombre_conjunto)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(300, height - 80, titulo)

    fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 115, f"Fecha de generación: {fecha_generacion}")
    c.drawString(50, height - 130, f"Generado por: {usuario}")

    y_position = height - 150
    data = [columnas] + datos

    table = Table(data, colWidths=[125] * len(columnas))
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 1, colors.black)
    ]))

    table.wrapOn(c, width, height)
    table.drawOn(c, 50, y_position - (len(datos) * 20))

    c.showPage()
    c.save()
    return filepath


def generar_excel(titulo, datos, columnas, filename, incluir_grafico=False):
    # Crear el directorio para guardar los reportes
    carpeta_reportes = os.path.abspath(os.path.join("static", "reportes"))
    os.makedirs(carpeta_reportes, exist_ok=True)

    # Crear el archivo Excel
    filepath = os.path.join(carpeta_reportes, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = titulo

    # Estilo del título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    titulo_celda = ws.cell(row=1, column=1)
    titulo_celda.value = titulo
    titulo_celda.font = Font(size=14, bold=True)
    titulo_celda.alignment = Alignment(horizontal="center")

    # Agregar encabezados
    for col_num, columna in enumerate(columnas, start=1):
        celda = ws.cell(row=2, column=col_num)
        celda.value = columna
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Agregar datos
    for row_num, fila in enumerate(datos, start=3):
        for col_num, valor in enumerate(fila, start=1):
            ws.cell(row=row_num, column=col_num).value = valor

    # Agregar gráfico si se solicita
    if incluir_grafico:
        chart = BarChart()
        chart.title = "Resumen de Datos"
        chart.style = 10
        chart.x_axis.title = columnas[0]
        chart.y_axis.title = "Valores"

        data = Reference(ws, min_col=2, min_row=2, max_col=len(columnas), max_row=len(datos) + 2)
        categorias = Reference(ws, min_col=1, min_row=3, max_row=len(datos) + 2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categorias)
        ws.add_chart(chart, f"E{len(datos) + 5}")

    # Guardar el archivo
    wb.save(filepath)
    return filepath


def generar_reportess():
    torres = Torre.query.all()
    apartamentos = Apartamento.query.all()
    roles = Roles.query.filter(Roles.id != 99).all()  
    if request.method == 'POST':

        tipo_reporte = request.form.get('tipo_reporte')
        print(tipo_reporte)
        #Reservas
        filtro_reservas = request.form.get('filtro_reservas')
        print(filtro_reservas)
        #PQRS
        filtro_pqrs = request.form.get('filtro_pqrs')
        tipo_pqrs = request.form.get('estado_pqrs')
        #Facturacion
        filtro_facturacion = request.form.get('filtro_facturacion')
        #Usuarios
        filtro_usuario_tipo = request.form.get('filtro_usuario_tipo')
        id_rol = request.form.get('rol_id')

        #Unidad Residencial
        id_torre  = request.form.get('torre_id')
        id_apartamento  = request.form.get('apartamento_id')

        #Fechas
        fecha_inicio  = request.form.get('fecha_inicio')
        fecha_fin  = request.form.get('fecha_fin')

# ------------------------------------------
        # RESERVAS APROBADAS
        if tipo_reporte == "reservas":
            if filtro_reservas =="confirmadas":
                titulo = "Reservas confirmadas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Horario", "Espacio"]
                query = Reserva.query.filter_by(id_estado=2)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename = f"reservas_aprobadas_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            # ------------------------------------------
            elif filtro_reservas == "pendientes":
                titulo = "Reservas en espera"
                columnas = ["Usuario","Unidad Residencial", "Fecha", "Horario", "Espacio", "Estado"]
                query = Reserva.query.filter_by(id_estado=1)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename = f"reservas_en_espera_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            # ------------------------------------------
            elif filtro_reservas == "canceladas":
                titulo = "Reservas Canceladas"
                columnas = ["Usuario","Unidad Residencial", "Fecha", "Horario", "Espacio", "Estado"]
                query = Reserva.query.filter_by(id_estado=3)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre  +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename = f"reservas_canceladas_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
                        # ------------------------------------------

        # ------------------------------------------
        elif tipo_reporte == "pqrs":
            if filtro_pqrs == "registradas":
                titulo = "PQRS registradas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=1, id_tipo= tipo_pqrs)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_registradas_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            # ------------------------------------------
            elif filtro_pqrs == "en_proceso":
                titulo = "PQRS en proceso"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=2, id_tipo= tipo_pqrs)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_en_proceso_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            # ------------------------------------------
            elif filtro_pqrs == "finalizadas":
                titulo = "PQRS Finalizadas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=3, id_tipo= tipo_pqrs)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_finalizadas_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))

                        # ------------------------------------------
            elif filtro_pqrs == "fuera":
                titulo = "PQRS Fuera de Tiempo"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=4, id_tipo= tipo_pqrs)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_fuera_tiempo_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
        # ----------------------------- -------------
        elif tipo_reporte == "facturacion":
            if filtro_facturacion == "en_mora":
                titulo = "Residentes en mora"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]
                query = Factura.query.filter_by(estado="Pendiente")
                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)

                datos = [[f.usuario_rel.nombre,f.usuario_rel.casa.torre.nombre  +" - "+  f.usuario_rel.casa.apartamento.numero, f.mes, f.year, f.total] for f in query.all()]
                filename = f"residentes_en_mora_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            # ------------------------------------------
            elif filtro_facturacion == "al_dia":
                titulo = "Residentes al día"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]
                query = Factura.query.filter_by(estado="Aprobado")

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)


                datos = [[f.usuario_rel.nombre,f.usuario_rel.casa.torre.nombre  +" - "+  f.usuario_rel.casa.apartamento.numero, f.mes, f.year, f.total] for f in query.all()]
                filename = f"residentes_al_dia_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            elif filtro_facturacion == "por_unidad":
                titulo = "Reporte por Residente"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)


                datos = [[f.usuario_rel.nombre,f.usuario.casa.torre.nombre  +" - "+  f.usuario.casa.apartamento.numero, f.mes, f.year, f.total] for f in Factura.query.all()]
                filename = f"residente_informe_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
        elif tipo_reporte == "usuarios":
            if filtro_usuario_tipo == "habilitados":
                titulo = "Reporte Usuarios Habilitados"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Estado"]
                query = Usuarios.query.filter_by(estado=1)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)


                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, "Habilitado"] for u in query.all()]
                filename = f"usuarios_habilitados_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            elif filtro_usuario_tipo == "inhabilitados":
                titulo = "Reporte Usuarios Inhabilitados"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Estado"]
                query = Usuarios.query.filter_by(estado=2)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)

                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, "Inhabilitado"] for u in query.all()]
                filename = f"usuarios_inhabilitados_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
            elif filtro_usuario_tipo == "por_rol":
                titulo = "Reporte Usuarios Por Rol"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Rol"]
                query = Usuarios.query.filter_by(id_rol==id_rol)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)


                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, u.rol.nombre] for u in query.all()]
                filename = f"usuarios_por_rol_{datetime.now().strftime('%Y%m%d')}.pdf"
                registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))
        else:
            flash("Seleccione un tipo de reporte válido.", "danger")
            return redirect(url_for('reportes.generar_reportes'))
 
        
        filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        return send_file(filepath, as_attachment=True)

    return render_template('reportes/reportes.html', torres=torres, apartamentos=apartamentos, roles=roles)


@reportes_bp.route('/reportes', methods=['GET', 'POST'])
@login_required
@require_permission('Generar Informes')
def generar_reportes():
    torres = Torre.query.all()
    apartamentos = Apartamento.query.all()
    roles = Roles.query.filter(Roles.id != 99).all()

    if request.method == 'POST':
        tipo_reporte = request.form.get('tipo_reporte')
        filtros = {
            'reservas': request.form.get('filtro_reservas'),
            'pqrs': request.form.get('filtro_pqrs'),
            'facturacion': request.form.get('filtro_facturacion'),
            'usuarios': request.form.get('filtro_usuario_tipo'),
        }
        id_rol = request.form.get('rol_id')
        id_torre = request.form.get('torre_id')
        id_apartamento = request.form.get('apartamento_id')
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        tipo_exportacion = request.form.get('export')

        # Mapeo de funciones de generación de reportes
        reportes = {
            'reservas': generar_reporte_reservas,
            'pqrs': generar_reporte_pqrs,
            'facturacion': generar_reporte_facturacion,
            'usuarios': generar_reporte_usuarios,
        }

        if tipo_reporte in reportes:
            return reportes[tipo_reporte](filtros[tipo_reporte], id_rol, id_torre, id_apartamento, fecha_inicio, fecha_fin, tipo_exportacion)
        else:
            flash("Seleccione un tipo de reporte válido.", "danger")
            return redirect(url_for('reportes.generar_reportes'))

    return render_template('reportes/reportes.html', torres=torres, apartamentos=apartamentos, roles=roles)


def generar_reporte_reservas(filtro, id_rol, id_torre, id_apartamento, fecha_inicio, fecha_fin, tipo_exportacion):
    estados = {
        'confirmadas': 2,
        'pendientes': 3,
        'canceladas': 4
    }
    estado = estados.get(filtro)
    if estado is None:
        flash("Filtro de reservas no válido.", "danger")
        return redirect(url_for('reportes.generar_reportes'))

    titulo = f"Reservas {filtro}"
    columnas = ["Usuario", "Unidad Residencial", "Fecha", "Horario", "Espacio"]
    query = Reserva.query.filter_by(id_estado=estado)

    query = aplicar_filtros_comunes(query, Reserva.usuario, id_torre, id_apartamento, Reserva.fecha, fecha_inicio, fecha_fin)

    datos = [
        [r.usuario.nombre, f"{r.usuario.casa.torre.nombre} - {r.usuario.casa.apartamento.numero}", r.fecha, r.horario, r.espacios.nombre]
        for r in query.all()
    ]

    if tipo_exportacion == 'excel':
        filename = f"reservas_{filtro}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {filename}")
        filepath = generar_excel(titulo, datos, columnas, filename, incluir_grafico=True)
        return send_file(filepath, as_attachment=True)
    else:
        filename = f"reservas_{filtro}_{datetime.now().strftime('%Y%m%d')}.pdf"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {titulo}")
        filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        return send_file(filepath, as_attachment=True)


def generar_reporte_pqrs(filtro, id_rol, id_torre, id_apartamento, fecha_inicio, fecha_fin, tipo_exportacion):
    estados = {
        'registradas': 1,
        'en_proceso': 2,
        'finalizadas': 3,
        'fuera': 4
    }
    estado = estados.get(filtro)
    if estado is None:
        flash("Filtro de PQRS no válido.", "danger")
        return redirect(url_for('reportes.generar_reportes'))

    tipo_pqrs = request.form.get('tipo_pqrs')
    titulo = f"PQRS {filtro.replace('_', ' ')}"
    columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
    query = PQRS.query.filter_by(id_estado=estado, id_tipo=tipo_pqrs)

    query = aplicar_filtros_comunes(query, PQRS.usuario, id_torre, id_apartamento, PQRS.fecha_creacion, fecha_inicio, fecha_fin)

    datos = [
        [p.usuario.nombre, f"{p.usuario.casa.torre.nombre} - {p.usuario.casa.apartamento.numero}", p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre]
        for p in query.all()
    ]

    if tipo_exportacion == 'excel':
        filename = f"pqrs_{filtro}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {filename}")
        filepath = generar_excel(titulo, datos, columnas, filename, incluir_grafico=True)
        return send_file(filepath, as_attachment=True)
    else:
        filename = f"pqrs_{filtro}_{datetime.now().strftime('%Y%m%d')}.pdf"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {titulo}")
        filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        return send_file(filepath, as_attachment=True)


def generar_reporte_facturacion(filtro, id_rol, id_torre, id_apartamento, fecha_inicio, fecha_fin, tipo_exportacion):
    estados = {
        'en_mora': "Pendiente",
        'al_dia': "Aprobado"
    }
    estado = estados.get(filtro)
    titulo = ""
    if estado:
        titulo = f"Residentes {'en mora' if filtro == 'en_mora' else 'al día'}"
        query = Factura.query.filter_by(estado=estado)
    elif filtro == "por_unidad":
        titulo = "Reporte por Residente"
        query = Factura.query
    else:
        flash("Filtro de facturación no válido.", "danger")
        return redirect(url_for('reportes.generar_reportes'))

    columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]
    query = aplicar_filtros_comunes(query, Factura.usuario_rel, id_torre, id_apartamento, Factura.fecha_emision, fecha_inicio, fecha_fin)

    datos = [
        [f.usuario_rel.nombre, f"{f.usuario_rel.casa.torre.nombre} - {f.usuario_rel.casa.apartamento.numero}", f.mes, f.year, f.total]
        for f in query.all()
    ]
    if tipo_exportacion == 'excel':
        filename = f"facturacion_{filtro}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {filename}")
        filepath = generar_excel(titulo, datos, columnas, filename, incluir_grafico=True)
        return send_file(filepath, as_attachment=True)
    else:
        filename = f"facturacion_{filtro}_{datetime.now().strftime('%Y%m%d')}.pdf"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {titulo}")
        filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        return send_file(filepath, as_attachment=True)


def generar_reporte_usuarios(filtro, id_rol, id_torre, id_apartamento, fecha_inicio, fecha_fin, tipo_exportacion):
    estados = {
        'habilitados': 1,
        'inhabilitados': 2
    }
    titulo = ""
    if filtro in estados:
        estado = estados[filtro]
        titulo = f"Usuarios {'Habilitados' if filtro == 'habilitados' else 'Inhabilitados'}"
        query = Usuarios.query.filter_by(estado=estado)
    elif filtro == "por_rol":
        titulo = "Usuarios por Rol"
        query = Usuarios.query.filter_by(id_rol=id_rol)
    else:
        flash("Filtro de usuarios no válido.", "danger")
        return redirect(url_for('reportes.generar_reportes'))

    columnas = ["Usuario", "Unidad Residencial", "Identificación", "Correo", "Teléfono", "Estado/Rol"]
    query = aplicar_filtros_comunes(query, Usuarios, id_torre, id_apartamento)

    for u in query.all():
        casa = u.casa
        if casa and casa.torre and casa.apartamento:
            unidad = f"{casa.torre.nombre} - {casa.apartamento.numero}"
        else:
            unidad = "Sin asignar"


    datos = [
        [u.nombre, unidad, u.identificacion, u.email, u.telefono, u.rol.nombre if filtro == "por_rol" else ("Habilitado" if u.estado == 1 else "Inhabilitado")]
        for u in query.all()
    ]
    if tipo_exportacion == 'excel':
        filename = f"usuarios_{filtro}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {filename}")
        filepath = generar_excel(titulo, datos, columnas, filename, incluir_grafico=True)
        return send_file(filepath, as_attachment=True)
    else:
        filename = f"usuarios_{filtro}_{datetime.now().strftime('%Y%m%d')}.pdf"
        registrar_log(current_user.id, "Reportes", f"Se Generó Reporte {titulo}")
        filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        return send_file(filepath, as_attachment=True)



def aplicar_filtros_comunes(query, modelo_usuario, id_torre=None, id_apartamento=None, campo_fecha=None, fecha_inicio=None, fecha_fin=None):
    if id_torre:
        query = query.join(modelo_usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
    if id_apartamento:
        query = query.join(modelo_usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
    if campo_fecha and fecha_inicio:
        query = query.filter(campo_fecha >= fecha_inicio)
    if campo_fecha and fecha_fin:
        query = query.filter(campo_fecha <= fecha_fin)
    return query
